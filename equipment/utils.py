import json
import logging
from datetime import datetime

import requests
from django.db.models import Q, F
from django.http import HttpResponse
from openpyxl import load_workbook
from io import BytesIO

from basics.models import WorkSchedulePlan
from equipment.models import EquipApplyOrder
from system.models import User, Section


def gen_template_response(export_fields_dict, data, file_name):
    export_fields = list(export_fields_dict.values())
    sheet_heads = list(export_fields_dict.keys())
    wb = load_workbook('xlsx_template/example.xlsx')
    ws = wb.worksheets[0]
    sheet = wb.copy_worksheet(ws)
    for idx, sheet_head in enumerate(sheet_heads):
        sheet.cell(1, idx + 1).value = sheet_head

    data_row = 2
    for i in data:
        for col_num, data_key in enumerate(export_fields):
            sheet.cell(data_row, col_num+1).value = i[data_key]
        data_row += 1

    wb.remove_sheet(ws)
    output = BytesIO()
    wb.save(output)
    # 重新定位到开始
    output.seek(0)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    filename = file_name
    response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
        'ISO-8859-1') + '.xls'
    response.write(output.getvalue())
    return response


class DinDinAPI(object):
    APP_KEY = 'dingpygx79tszcicv6jb'
    APP_SECRET = 'LBTg3wzgkDZ5b2l5pqRBrUy8XY8npiaPKYqBk5fCfCKMg_PEB7FJFpYMj8OfiiUu'
    AGENT_ID = '1336171749'

    def __init__(self):
        self.access_token = self.get_token()

    def get_token(self):
        """
            获取钉钉授权认证token
        @return:
        """
        url = 'https://oapi.dingtalk.com/gettoken'
        ret = requests.get(url, params={'appkey': self.APP_KEY, 'appsecret': self.APP_SECRET})
        data = json.loads(ret.text)
        return data.get('access_token')

    def get_user_id(self, phone_num):
        """
            根据手机号获取钉钉用户id
        @param phone_num:
        """
        url = 'https://oapi.dingtalk.com/topapi/v2/user/getbymobile'
        ret = requests.post(url, params={'access_token': self.access_token}, data={"mobile": phone_num})
        data = json.loads(ret.text)
        return data.get('result').get('userid') if data.get('errcode') == 0 else ''

    def get_user_attendance(self, user_ids, begin_time=datetime.now().date(), end_time=datetime.now().date()):
        """
            获取用户考勤信息
        @param user_ids: 用户id列表
        @param begin_time: 开始工作日
        @param end_time: 结束工作日
        @return: 考勤打卡列表
        """
        url = 'https://oapi.dingtalk.com/attendance/list'
        data = {
            "workDateFrom": str(begin_time) + ' 00:00:00',
            "offset": 0,
            "userIdList": user_ids,
            "limit": 10,
            "isI18n": False,
            "workDateTo": str(end_time) + ' 00:00:00'
        }
        ret = requests.post(url, params={'access_token': self.access_token}, json=data)
        data = json.loads(ret.text)
        return data.get('recordresult')

    def send_message(self, user_ids, content):
        """
            发送钉钉工作消息给用户
        @param user_ids:
        """
        url = 'https://oapi.dingtalk.com/topapi/message/corpconversation/asyncsend_v2'
        data = {
            "msg": {
                "msgtype": "oa",
                "oa": {
                    "message_url": "http://dingtalk.com",
                    "head": {
                        "bgcolor": "FFBBBBBB",
                        "text": "头部标题"
                    },
                    "body": content
                }
            },
            "agent_id": self.AGENT_ID,
            "userid_list": ",".join(user_ids),
            "to_all_user": False
        }
        ret = requests.post(url, params={'access_token': self.access_token}, json=data)
        data = json.loads(ret.text)
        if not data.get('errcode') == 0:
            print('请求错误')


if __name__ == '__main__':
    a = DinDinAPI()
    # print(a.get_user_id('15607115901'))
    # # print(a.get_user_attendance(['0206046007692894']))
    # # a.send_message(['0206046007692894'])


def get_staff_status(ding_api, section_name='维修部', group=''):
    """section_name: 部门"""
    result = []
    filter_kwargs = {'section_users__repair_group': group} if group else {'section_users__repair_group__isnull': False}
    # 获取部门所有员工信息
    staffs = Section.objects.filter(Q(name=section_name) | Q(parent_section__name=section_name), **filter_kwargs)\
        .annotate(username=F('section_users__username'), phone_number=F('section_users__phone_number'),
                  group=F('section_users__repair_group'),
                  uid=F('section_users__id'), leader=F('in_charge_user__username'),
                  leader_phone_number=F('in_charge_user__phone_number'))\
        .values('username', 'phone_number', 'uid', 'leader', 'leader_phone_number', 'group')
    staffs_detail = [i for i in staffs if i['group'] == group] if group else staffs
    for staff in staffs_detail:
        staff_dict = {'id': staff.get('uid'), 'phone_number': staff.get('phone_number'), 'optional': False,
                      'username': staff.get('username'), 'group': staff.get('group'), 'leader': staff.get('leader'),
                      'leader_phone_number': staff.get('leader_phone_number')}
        # 根据手机号获取用户钉钉uid
        ding_uid = ding_api.get_user_id(staff.get('phone_number'))
        if ding_uid:
            # # 查询考勤记录
            staff_dict['ding_uid'] = ding_uid
            # records = ding_api.get_user_attendance([ding_uid])
            # if records and not len([i for i in records if i['checkType'] != 'OnDuty']):
            #     staff_dict['optional'] = True
            staff_dict['optional'] = True
        result.append(staff_dict)
    return result


def get_ding_uids(ding_api, pks=None, names=None):
    """pks：维修单id列表"""
    user_ids, assign_user_list = [], []
    if pks:
        assign_user_list = EquipApplyOrder.objects.filter(id__in=pks).values_list('assign_user').distinct()
    if names:
        assign_user_list = names
    phone_numbers = User.objects.filter(username__in=assign_user_list).values_list('phone_number')
    for phone in phone_numbers:
        res = ding_api.get_user_id(phone)
        if res:
            user_ids.append(res)
    return user_ids


class AutoDispatch(object):
    """自动派单"""
    def __init__(self):
        # self.applys = EquipApplyOrder.objects.filter(status='已生成')
        # self.inspections = EquipInspectionOrder.objects.filter(status='已生成')
        self.logger = logging.getLogger('send_order')

    def send_order(self, order, choice_all_user):
        # 班组
        group = self.get_group_info()
        # 维修部门下改班组人员
        choice_all_user = get_staff_status(DinDinAPI(), group=group)
        working_persons = [i for i in choice_all_user if i['optional']]
        if not working_persons:
            # 发送消息给上级
            pass
        for per in working_persons:
            new_applyed = EquipApplyOrder.objects.filter(status='已生成').first()
            if not new_applyed:
                self.logger.info('没有新生成的维修单')
                break
            processing_order = EquipApplyOrder.objects.filter(Q(result_repair_final_result='等待'), status='已开始',
                                                              repair_user=per['username'])
            if processing_order:
                continue


        # 分派维修单
        # 派单成功发送钉钉消息上级和当班人员, 失败发送给上级
        pass

    def get_group_info(self):
        now_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        group = '早班' if '08:00:00' < now_date[:10] < '20:00:00' else '夜班'
        record = WorkSchedulePlan.objects.filter(plan_schedule__day_time=now_date[11:], classes__global_name=group,
                                                 plan_schedule__work_schedule__work_procedure__global_name='密炼').first()
        return record.group.global_name


if __name__ == '__main__':
    auto_dispatch = AutoDispatch()
    auto_dispatch.send_order()

